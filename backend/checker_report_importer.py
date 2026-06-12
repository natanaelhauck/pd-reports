import hashlib
import os
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from course_checker import (
    CourseCheckerError,
    course_name_is_excluded_from_consumption,
    expand_student_courses_to_official,
    load_existing_consumption_enrichment,
    normalize_student_result,
    parse_boolish,
    resolve_official_course_catalog,
    normalize_email,
    parse_percent_input,
    parse_total_certifiable,
    persist_consumption_run,
    resolve_student_name,
    text,
)
from course_rules import COURSE_CONSUMPTION_TOTAL_CERTIFIABLE


BACKEND_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BACKEND_DIR.parent
DEFAULT_REPORT_PATH = "checker/relatorio_final.xlsx"
REQUIRED_SHEETS = ("Resumo por aluno", "Cursos por aluno")
SOURCE_TYPE_CHECKER_REPORT_XLSX = "checker_report_xlsx"


def resolve_project_path(value):
    path = Path(value or "")
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path


def load_report_path_from_env(env=None):
    env = env or os.environ
    report_path = env.get("CHECKER_REPORT_XLSX_PATH", DEFAULT_REPORT_PATH)
    return resolve_project_path(report_path)


def ensure_report_exists(report_path):
    resolved = Path(report_path)
    if not resolved.is_file():
        raise CourseCheckerError(f"Relatorio do checker nao encontrado: {resolved}")
    return resolved


def ensure_report_is_valid(report_path):
    resolved = ensure_report_exists(report_path)
    workbook = load_workbook(resolved, read_only=True, data_only=True)
    try:
        missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook.sheetnames]
        if missing:
            raise CourseCheckerError(
                "O relatorio precisa conter as abas 'Resumo por aluno' e 'Cursos por aluno'."
            )
        return resolved
    finally:
        workbook.close()


def sha256_file(path):
    hasher = hashlib.sha256()
    with Path(path).open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def header_index(headers):
    return {text(header): index for index, header in enumerate(headers)}


def pick_value(row, indices, *candidates):
    for candidate in candidates:
        index = indices.get(text(candidate))
        if index is not None and index < len(row):
            value = row[index]
            if value is not None and value != "":
                return value
    return None


def load_report_rows(report_path):
    resolved = ensure_report_is_valid(report_path)
    workbook = load_workbook(resolved, read_only=True, data_only=True)
    try:
        resumo_ws = workbook[REQUIRED_SHEETS[0]]
        cursos_ws = workbook[REQUIRED_SHEETS[1]]

        resumo_rows = list(resumo_ws.iter_rows(values_only=True))
        cursos_rows = list(cursos_ws.iter_rows(values_only=True))
        if len(resumo_rows) < 2 or len(cursos_rows) < 2:
            raise CourseCheckerError("O relatorio do checker esta vazio ou incompleto.")

        resumo_headers = header_index(resumo_rows[0])
        cursos_headers = header_index(cursos_rows[0])

        resumo = {}
        resumo_duplicados = 0
        for row in resumo_rows[1:]:
            email = normalize_email(pick_value(row, resumo_headers, "Email"))
            if not email:
                continue
            if email in resumo:
                resumo_duplicados += 1
                continue
            resumo[email] = {
                "email": email,
                "nome": text(pick_value(row, resumo_headers, "Aluno")) or email,
            }

        cursos_por_email = defaultdict(list)
        catalog_by_course = {}
        cursos_ignorados = 0
        cursos_invalidos = 0
        for ordem, row in enumerate(cursos_rows[1:], start=1):
            email = normalize_email(pick_value(row, cursos_headers, "Email"))
            if not email:
                cursos_invalidos += 1
                continue

            course_id = text(pick_value(row, cursos_headers, "Course ID", "CourseId", "course_id"))
            course_name = text(pick_value(row, cursos_headers, "Curso", "Course", "course_name"))
            if not course_id:
                course_id = course_name
            if not course_name:
                course_name = course_id
            if not course_name:
                cursos_invalidos += 1
                continue
            if course_name_is_excluded_from_consumption(course_name) or course_name_is_excluded_from_consumption(course_id):
                cursos_ignorados += 1
                catalog_by_course.setdefault(course_id, {
                    "courseId": course_id,
                    "courseName": course_name,
                    "certificavel": False,
                    "ignored": True,
                    "ordem": ordem,
                })
                continue

            percentual_100, status_scale = parse_percent_input(pick_value(row, cursos_headers, "Percentual"))
            status = text(pick_value(row, cursos_headers, "Status")) or (
                "Concluído" if status_scale >= 0.6 else "Em andamento" if status_scale > 0 else "Não iniciado"
            )
            certificado_gerado = parse_boolish(pick_value(row, cursos_headers, "Certificado gerado"))

            catalog_by_course.setdefault(course_id, {
                "courseId": course_id,
                "courseName": course_name,
                "certificavel": True,
                "ignored": False,
                "ordem": ordem,
            })
            cursos_por_email[email].append({
                "courseId": course_id,
                "courseName": course_name,
                "status": status,
                "percentual": percentual_100,
                "certificadoGerado": certificado_gerado,
            })

        total_linhas_cursos = len(cursos_rows) - 1
        return {
            "resumo": resumo,
            "cursos_por_email": cursos_por_email,
            "catalogo": list(catalog_by_course.values()),
            "warnings": {
                "resumo_duplicados": resumo_duplicados,
                "cursos_ignorados": cursos_ignorados,
                "cursos_invalidos": cursos_invalidos,
                "total_linhas_cursos": total_linhas_cursos,
            },
        }
    finally:
        workbook.close()


def build_payload_from_report(report_path, enrichment_by_email=None, total_certifiable=None, source_type=SOURCE_TYPE_CHECKER_REPORT_XLSX, original_filename=None):
    total_certifiable = parse_total_certifiable(
        total_certifiable or COURSE_CONSUMPTION_TOTAL_CERTIFIABLE
    )
    report = load_report_rows(report_path)
    enrichment_by_email = enrichment_by_email or {}
    official_catalog = resolve_official_course_catalog(
        env=os.environ,
        expected_total=total_certifiable,
    )
    students = []
    emails = set(report["resumo"].keys()) | set(report["cursos_por_email"].keys())
    for email in sorted(emails):
        base = report["resumo"].get(email, {"email": email, "nome": email})
        consumo_enrichment = enrichment_by_email.get(email) or {}
        checker_name = base.get("nome") or email
        consumo_name = consumo_enrichment.get("nome") or ""
        student = {
            "email": email,
            "nomeChecker": checker_name,
            "nomeConsumo": consumo_name,
            "nome": resolve_student_name(
                pd_name=None,
                consumption_name=consumo_name,
                checker_name=checker_name,
                fallback_identifier=email,
            ),
            "desafioFinal": False,
            "ingresso": None,
            "cursos": report["cursos_por_email"].get(email, []),
        }
        enrich = consumo_enrichment
        student["desafioFinal"] = bool(enrich.get("desafioFinal", False))
        student["ingresso"] = enrich.get("ingresso")
        student["cursos"] = expand_student_courses_to_official(
            student.get("cursos", []),
            official_catalog,
            username=email,
        )
        normalize_student_result(student, total_certifiable)
        students.append(student)

    payload = {
        "students": students,
        "warnings": [],
        "courseCatalog": official_catalog,
        "sourceFilesInfo": build_source_files_info(
            report_path,
            payload_students=students,
            payload_courses=sum(len(student.get("cursos", [])) for student in students),
            source_type=source_type,
            original_filename=original_filename,
            formula=f"consumoPercentual = round(sum(percentual dos {total_certifiable} cursos certificáveis) / {total_certifiable}, 2)",
        ),
        "totals": {
            "students": len(students),
            "courses": sum(len(student.get("cursos", [])) for student in students),
            "resumoDuplicados": report["warnings"]["resumo_duplicados"],
            "cursosIgnorados": report["warnings"]["cursos_ignorados"],
            "cursosInvalidos": report["warnings"]["cursos_invalidos"],
            "linhasCursos": report["warnings"]["total_linhas_cursos"],
        },
        "source": source_type,
        "sourceType": source_type,
        "totalCertifiable": total_certifiable,
    }

    if report["warnings"]["resumo_duplicados"]:
        payload["warnings"].append(
            f"{report['warnings']['resumo_duplicados']} e-mails duplicados foram ignorados no resumo do relatório."
        )
    if report["warnings"]["cursos_ignorados"]:
        payload["warnings"].append(
            f"{report['warnings']['cursos_ignorados']} linhas de cursos foram ignoradas por regra oficial."
        )
    if report["warnings"]["cursos_invalidos"]:
        payload["warnings"].append(
            f"{report['warnings']['cursos_invalidos']} linhas de cursos inválidas foram descartadas."
        )

    return payload


def build_source_files_info(report_path, payload_students=None, payload_courses=None, source_type=SOURCE_TYPE_CHECKER_REPORT_XLSX, original_filename=None, formula="consumoPercentual = round(sum(percentual dos cursos certificáveis) / 22, 2)"):
    resolved = ensure_report_exists(report_path)
    stat = resolved.stat()
    return {
        "sourceType": source_type,
        "arquivoOriginal": original_filename or resolved.name,
        "tamanhoBytes": stat.st_size,
        "sha256": sha256_file(resolved),
        "quantidadeAlunos": len(payload_students or []),
        "quantidadeCursos": int(payload_courses) if payload_courses is not None else sum(len(student.get("cursos", [])) for student in (payload_students or [])),
        "formula": formula,
        "abas": list(REQUIRED_SHEETS),
    }


def importar_relatorio_checker_xlsx(
    conn,
    report_path,
    *,
    triggered_by_user_id=None,
    total_certifiable=None,
    source_type=SOURCE_TYPE_CHECKER_REPORT_XLSX,
    enrichment_by_email=None,
    original_filename=None,
):
    total_certifiable = parse_total_certifiable(
        total_certifiable or COURSE_CONSUMPTION_TOTAL_CERTIFIABLE
    )
    if enrichment_by_email is None:
        try:
            enrichment_by_email = load_existing_consumption_enrichment(env=os.environ)
        except Exception:
            enrichment_by_email = {}
    payload = build_payload_from_report(
        report_path,
        enrichment_by_email=enrichment_by_email,
        total_certifiable=total_certifiable,
        source_type=source_type,
        original_filename=original_filename,
    )
    return persist_consumption_run(
        conn,
        payload,
        triggered_by_user_id=triggered_by_user_id,
        source_files_info=payload.get("sourceFilesInfo"),
        source_type=source_type,
    )
