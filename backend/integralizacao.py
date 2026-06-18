import copy
import json
import math
import os
import re
import time
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

import psycopg2
from openpyxl import load_workbook

from course_rules import (
    COURSE_CONSUMPTION_TOTAL_CERTIFIABLE,
    course_name_is_excluded_from_consumption,
    missing_certificate_course_sort_key,
    official_course_sort_key,
)
from course_checker import (
    CourseCheckerError,
    expand_student_courses_to_official,
    resolve_official_course_catalog,
)
from consumption_repository import (
    get_consumption_courses_from_run_by_email,
    get_consumption_courses_from_run,
    get_consumption_student_from_run_by_email,
    get_consumption_students_from_run,
    get_latest_successful_run,
)


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX_PATH = "dados/consumo_local.xlsx"
DEFAULT_SHEET_NAME = "Resultado"
DEFAULT_HORAS_TOTAIS = 154
DEFAULT_PRAZO_FINAL = "2026-11-30"
DEFAULT_CACHE_TTL_SECONDS = 60
DEFAULT_CONSUMPTION_SOURCE_MODE = "auto"
DEFAULT_TOTAL_CURSOS_CERTIFICAVEIS = COURSE_CONSUMPTION_TOTAL_CERTIFIABLE
TOTAL_CURSOS_CERTIFICAVEIS = DEFAULT_TOTAL_CURSOS_CERTIFICAVEIS
MIN_MINUTOS_DIA = 30
MIN_HORAS_DIA = MIN_MINUTOS_DIA / 60
EXCEL_EPOCH = datetime(1899, 12, 30)

_CACHE = {
    "key": None,
    "expires_at": 0,
    "data": None,
}


class IntegralizacaoError(Exception):
    pass


class IntegralizacaoArquivoNaoEncontrado(IntegralizacaoError):
    pass


class IntegralizacaoConfigInvalida(IntegralizacaoError):
    pass


class IntegralizacaoPlanilhaInvalida(IntegralizacaoError):
    pass


class IntegralizacaoNeonSemDados(IntegralizacaoError):
    pass


class IntegralizacaoNeonIndisponivel(IntegralizacaoError):
    pass


def limpar_cache_integralizacao():
    _CACHE["key"] = None
    _CACHE["expires_at"] = 0
    _CACHE["data"] = None


def normalizar_email(valor):
    return str(valor or "").strip().lower()


def sem_acentos(valor):
    return "".join(
        ch for ch in unicodedata.normalize("NFD", str(valor or ""))
        if unicodedata.category(ch) != "Mn"
    )


def chave_campo(valor):
    return re.sub(r"[^a-z0-9]+", " ", sem_acentos(valor).lower()).strip()


def texto(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def parse_numero(valor):
    if valor is None or valor == "":
        return 0
    if isinstance(valor, (int, float)):
        numero = float(valor)
    else:
        numero = float(str(valor).strip().replace(",", "."))
    if not math.isfinite(numero):
        return 0
    return round(numero, 2)


def numero_seguro(valor):
    try:
        return parse_numero(valor)
    except (TypeError, ValueError):
        return 0


def inteiro_seguro(valor, padrao=0):
    try:
        if valor is None or valor == "":
            return padrao
        return int(valor)
    except (TypeError, ValueError):
        return padrao


def clamp(valor, minimo=0, maximo=100):
    return max(minimo, min(maximo, valor))


def parse_bool_certificado(valor):
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, (int, float)):
        return valor > 0
    normalizado = sem_acentos(valor).strip().lower()
    return normalizado in {"sim", "true", "1", "yes", "gerado", "certificado"}


def parse_percentual(valor):
    numero = numero_seguro(valor)
    if 0 < numero <= 1:
        numero *= 100
    return round(clamp(numero), 2)


def parse_data_excel(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, (int, float)) and valor:
        return (EXCEL_EPOCH + timedelta(days=int(valor))).date()

    raw = texto(valor)
    if not raw:
        return None
    if re.fullmatch(r"\d+(\.\d+)?", raw):
        return (EXCEL_EPOCH + timedelta(days=int(float(raw)))).date()
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw[:10], formato).date()
        except ValueError:
            pass
    return None


def formatar_data_br(valor):
    data = parse_data_excel(valor)
    if not data:
        return texto(valor)
    return data.strftime("%d/%m/%Y")


def formatar_data_iso(data_valor):
    if not data_valor:
        return ""
    return data_valor.isoformat()


def formatar_data_br_date(data_valor):
    if not data_valor:
        return ""
    return data_valor.strftime("%d/%m/%Y")


def parse_prazo_final(valor):
    raw = texto(valor or DEFAULT_PRAZO_FINAL)
    try:
        return date.fromisoformat(raw)
    except ValueError as exc:
        raise IntegralizacaoConfigInvalida(
            "INTEGRALIZACAO_PRAZO_FINAL deve estar no formato YYYY-MM-DD."
        ) from exc


def parse_horas_totais(valor):
    horas = numero_seguro(valor or DEFAULT_HORAS_TOTAIS)
    if horas <= 0:
        raise IntegralizacaoConfigInvalida("INTEGRALIZACAO_HORAS_TOTAIS deve ser maior que zero.")
    return horas


def parse_total_cursos_certificaveis(valor):
    try:
        total = int(valor or DEFAULT_TOTAL_CURSOS_CERTIFICAVEIS)
    except (TypeError, ValueError) as exc:
        raise IntegralizacaoConfigInvalida(
            "COURSE_CONSUMPTION_TOTAL_CERTIFIABLE deve ser um numero inteiro."
        ) from exc
    if total <= 0:
        raise IntegralizacaoConfigInvalida(
            "COURSE_CONSUMPTION_TOTAL_CERTIFIABLE deve ser maior que zero."
        )
    return total


def parse_consumption_source_mode(valor):
    modo = texto(valor or DEFAULT_CONSUMPTION_SOURCE_MODE).lower()
    if modo not in {"xlsx", "neon", "auto"}:
        raise IntegralizacaoConfigInvalida(
            "CONSUMPTION_SOURCE_MODE deve ser xlsx, neon ou auto."
        )
    return modo


def resolver_caminho(caminho):
    caminho = texto(caminho or DEFAULT_XLSX_PATH)
    path = Path(caminho)
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path


def config_integralizacao(env=None, **overrides):
    env = env or os.environ
    caminho_configurado = overrides.get(
        "xlsx_path",
        env.get("INTEGRALIZACAO_XLSX_PATH", DEFAULT_XLSX_PATH),
    )
    sheet_name = texto(overrides.get(
        "sheet_name",
        env.get("INTEGRALIZACAO_SHEET_NAME", DEFAULT_SHEET_NAME),
    )) or DEFAULT_SHEET_NAME
    horas_totais = parse_horas_totais(overrides.get(
        "horas_totais",
        env.get("INTEGRALIZACAO_HORAS_TOTAIS", DEFAULT_HORAS_TOTAIS),
    ))
    prazo_final = parse_prazo_final(overrides.get(
        "prazo_final",
        env.get("INTEGRALIZACAO_PRAZO_FINAL", DEFAULT_PRAZO_FINAL),
    ))
    try:
        cache_ttl = int(overrides.get(
            "cache_ttl",
            env.get("INTEGRALIZACAO_CACHE_TTL_SECONDS", DEFAULT_CACHE_TTL_SECONDS),
        ))
    except (TypeError, ValueError) as exc:
        raise IntegralizacaoConfigInvalida(
            "INTEGRALIZACAO_CACHE_TTL_SECONDS deve ser um número inteiro."
        ) from exc
    source_mode = parse_consumption_source_mode(overrides.get(
        "source_mode",
        env.get("CONSUMPTION_SOURCE_MODE", DEFAULT_CONSUMPTION_SOURCE_MODE),
    ))
    total_cursos_certificaveis = parse_total_cursos_certificaveis(overrides.get(
        "total_cursos_certificaveis",
        env.get("COURSE_CONSUMPTION_TOTAL_CERTIFIABLE", DEFAULT_TOTAL_CURSOS_CERTIFICAVEIS),
    ))
    return {
        "xlsx_path_configurado": texto(caminho_configurado or DEFAULT_XLSX_PATH),
        "xlsx_path": resolver_caminho(caminho_configurado),
        "sheet_name": sheet_name,
        "horas_totais": horas_totais,
        "prazo_final": prazo_final,
        "cache_ttl": max(0, cache_ttl),
        "source": source_mode,
        "source_mode": source_mode,
        "database_url": texto(overrides.get("database_url", env.get("DATABASE_URL", ""))),
        "total_cursos_certificaveis": total_cursos_certificaveis,
    }


def cache_key(config, hoje, incluir_cursos=True):
    hoje_key = hoje.isoformat() if hoje else date.today().isoformat()
    return (
        str(config["xlsx_path"]),
        config["sheet_name"],
        config["horas_totais"],
        config["prazo_final"].isoformat(),
        config["source_mode"],
        bool(config.get("database_url")),
        config["total_cursos_certificaveis"],
        bool(incluir_cursos),
        hoje_key,
    )


def header_index(headers):
    return {chave_campo(header): index for index, header in enumerate(headers)}


def localizar_coluna(indices, *nomes):
    for nome in nomes:
        indice = indices.get(chave_campo(nome))
        if indice is not None:
            return indice
    return None


def valor_linha(row, indice):
    if indice is None or indice >= len(row):
        return None
    return row[indice]


def limpar_curso(raw):
    if not isinstance(raw, dict):
        return None
    curso = texto(raw.get("curso"))
    course_id = texto(raw.get("courseId") or raw.get("course_id") or raw.get("id"))
    status = texto(raw.get("status"))
    percentual = parse_percentual(raw.get("percentual"))
    certificado_gerado = parse_bool_certificado(
        raw.get("certificadoGerado", raw.get("certificado_gerado"))
    )
    if not curso and not course_id:
        return None
    return {
        "curso": curso or course_id,
        "courseId": course_id or curso,
        "status": status or "Não informado",
        "percentual": percentual,
        "certificadoGerado": certificado_gerado,
    }


def curso_removido_certificados(curso):
    nome = curso.get("curso") if isinstance(curso, dict) else curso
    return course_name_is_excluded_from_consumption(nome)


def remover_curso_de_lista(valor):
    return "; ".join(
        item
        for item in (texto(parte) for parte in texto(valor).split(";"))
        if item and not course_name_is_excluded_from_consumption(item)
    )


def curso_concluido(curso):
    status = chave_campo(curso.get("status"))
    return status == "concluido" or curso.get("percentual", 0) >= 100


def curso_em_andamento(curso):
    return chave_campo(curso.get("status")) == "em andamento"


def curso_nao_iniciado(curso):
    return chave_campo(curso.get("status")) == "nao iniciado"


def parse_cursos_json(valor):
    raw = texto(valor)
    if not raw:
        return []
    try:
        dados = json.loads(raw)
    except (TypeError, json.JSONDecodeError):
        return []
    if not isinstance(dados, list):
        return []
    cursos = []
    for item in dados:
        curso = limpar_curso(item)
        if curso and not curso_removido_certificados(curso):
            cursos.append(curso)
    return cursos


def carregar_catalogo_oficial_seguro(total_cursos_certificaveis):
    env = os.environ.copy()
    env.setdefault(
        "COURSE_CHECKER_CATALOG_PATH",
        str(PROJECT_ROOT / "checker" / "cursos_new.json"),
    )
    env.setdefault(
        "COURSE_CHECKER_IGNORE_PATH",
        str(PROJECT_ROOT / "checker" / "ignore_courses.json"),
    )
    try:
        return resolve_official_course_catalog(
            env=env,
            expected_total=total_cursos_certificaveis,
        )
    except CourseCheckerError:
        return []


def curso_para_certificados(curso):
    nome = texto(curso.get("courseName") or curso.get("curso") or curso.get("courseId"))
    course_id = texto(curso.get("courseId") or nome)
    return {
        "curso": nome or course_id,
        "courseId": course_id,
        "status": texto(curso.get("status")) or "Nao iniciado",
        "percentual": parse_percentual(curso.get("percentual")),
        "certificadoGerado": bool(curso.get("certificadoGerado")),
    }


def expandir_cursos_certificados(cursos, total_cursos_certificaveis):
    cursos = [curso for curso in (cursos or []) if curso and not curso_removido_certificados(curso)]
    catalogo = carregar_catalogo_oficial_seguro(total_cursos_certificaveis)
    if not catalogo:
        return sorted(cursos, key=official_course_sort_key)

    expandidos = expand_student_courses_to_official(cursos, catalogo)
    return [curso_para_certificados(curso) for curso in expandidos]


def agrupar_cursos(cursos):
    concluidos = []
    em_andamento = []
    nao_iniciados = []
    com_certificado = []
    sem_certificado = []

    for curso in cursos:
        if curso.get("certificadoGerado"):
            com_certificado.append(curso)
        else:
            sem_certificado.append(curso)

        if curso_concluido(curso):
            concluidos.append(curso)
        elif curso_em_andamento(curso):
            em_andamento.append(curso)
        elif curso_nao_iniciado(curso):
            nao_iniciados.append(curso)

    return {
        "concluidos": sorted(concluidos, key=official_course_sort_key),
        "emAndamento": sorted(em_andamento, key=official_course_sort_key),
        "naoIniciados": sorted(nao_iniciados, key=official_course_sort_key),
        "comCertificado": sorted(com_certificado, key=official_course_sort_key),
        "semCertificado": sorted(sem_certificado, key=missing_certificate_course_sort_key),
    }


def build_certificados(
    cells,
    desafio_final=False,
    total_cursos_certificaveis=TOTAL_CURSOS_CERTIFICAVEIS,
):
    cursos = expandir_cursos_certificados(
        parse_cursos_json(cells.get("cursosDetalhesJson")),
        total_cursos_certificaveis,
    )
    if desafio_final:
        grupos = agrupar_cursos(cursos)
        com_certificado = grupos["comCertificado"]
        return {
            "totalCursosCertificaveis": total_cursos_certificaveis,
            "cursosConcluidos": total_cursos_certificaveis,
            "certificadosGerados": total_cursos_certificaveis,
            "cursosEmAndamento": 0,
            "cursosNaoIniciados": 0,
            "cursosComCertificado": texto_cursos_resumo(com_certificado),
            "cursosSemCertificado": "",
            "cursos": cursos,
            "grupos": grupos,
        }

    grupos = agrupar_cursos(cursos)
    return {
        "totalCursosCertificaveis": total_cursos_certificaveis,
        "cursosConcluidos": len(grupos["concluidos"]),
        "certificadosGerados": len(grupos["comCertificado"]),
        "cursosEmAndamento": len(grupos["emAndamento"]),
        "cursosNaoIniciados": len(grupos["naoIniciados"]),
        "cursosComCertificado": remover_curso_de_lista(cells.get("cursosComCertificado")),
        "cursosSemCertificado": remover_curso_de_lista(cells.get("cursosSemCertificado")),
        "cursos": cursos,
        "grupos": grupos,
    }


def is_business_day(data_valor):
    return data_valor.weekday() < 5


def count_business_days(inicio, fim):
    if inicio > fim:
        return 0
    total = 0
    atual = inicio
    while atual <= fim:
        if is_business_day(atual):
            total += 1
        atual += timedelta(days=1)
    return total


def nth_business_day(inicio, numero):
    atual = inicio
    contador = 0
    while True:
        if is_business_day(atual):
            contador += 1
            if contador >= numero:
                return atual
        atual += timedelta(days=1)


def calcular_meta_diaria(horas_vistas, horas_totais, prazo_final, desafio_final=False, hoje=None):
    hoje = hoje or date.today()
    horas_restantes = round(max(0, horas_totais - horas_vistas), 2)
    percentual = 100 if desafio_final else round(clamp((horas_vistas / horas_totais) * 100), 2)
    aluno_concluido = bool(desafio_final or percentual >= 100 or horas_restantes <= 0)

    base = {
        "aplicavel": False,
        "mensagem": "",
        "prazoFinal": prazo_final.isoformat(),
        "prazoFinalFormatado": formatar_data_br_date(prazo_final),
        "horasRestantesCurso": 0 if aluno_concluido else horas_restantes,
        "minMinutosPorDia": MIN_MINUTOS_DIA,
        "diasUteisRestantes": 0,
        "semanasRestantes": 0,
        "horasPorDia": 0,
        "diasUteisNecessarios": 0,
        "dataPrevistaConclusao": "",
        "dataPrevistaConclusaoFormatada": "",
        "concluiAntesPrazo": False,
        "semana": {
            "segunda": 0,
            "terca": 0,
            "quarta": 0,
            "quinta": 0,
            "sexta": 0,
            "sabado": None,
        },
    }

    if desafio_final:
        base["mensagem"] = "Aluno concluiu o curso pelo Desafio Final."
        base["dataPrevistaConclusao"] = hoje.isoformat()
        base["dataPrevistaConclusaoFormatada"] = formatar_data_br_date(hoje)
        return base

    if aluno_concluido:
        base["mensagem"] = "Aluno já atingiu 100% de consumo."
        base["dataPrevistaConclusao"] = hoje.isoformat()
        base["dataPrevistaConclusaoFormatada"] = formatar_data_br_date(hoje)
        return base

    if prazo_final < hoje:
        base["mensagem"] = "Prazo final de consumo já passou."
        return base

    dias_uteis = count_business_days(hoje, prazo_final)
    if dias_uteis <= 0:
        base["mensagem"] = "Não há dias úteis disponíveis até o prazo final."
        return base

    media_para_prazo = horas_restantes / dias_uteis
    horas_por_dia = round(max(MIN_HORAS_DIA, media_para_prazo), 2)
    dias_necessarios = min(dias_uteis, math.ceil(horas_restantes / horas_por_dia))
    data_conclusao = nth_business_day(hoje, dias_necessarios)

    base.update({
        "aplicavel": True,
        "mensagem": "",
        "diasUteisRestantes": dias_uteis,
        "semanasRestantes": max(1, math.ceil(dias_uteis / 5)),
        "horasPorDia": horas_por_dia,
        "diasUteisNecessarios": dias_necessarios,
        "dataPrevistaConclusao": data_conclusao.isoformat(),
        "dataPrevistaConclusaoFormatada": formatar_data_br_date(data_conclusao),
        "concluiAntesPrazo": dias_necessarios < dias_uteis,
        "semana": {
            "segunda": horas_por_dia,
            "terca": horas_por_dia,
            "quarta": horas_por_dia,
            "quinta": horas_por_dia,
            "sexta": horas_por_dia,
            "sabado": None,
        },
    })
    return base


def percentual_integralizacao(horas_vistas, horas_totais, desafio_final=False):
    if desafio_final:
        return 100
    return round(clamp((horas_vistas / horas_totais) * 100), 2)


def montar_aluno_integralizacao(cells, config, hoje=None):
    email = texto(cells.get("email"))
    email_normalizado = normalizar_email(email)
    if not email_normalizado:
        return None

    horas_vistas = numero_seguro(cells.get("horasVistas"))
    data_entrada = parse_data_excel(cells.get("dataEntrada"))
    desafio_final = chave_campo(cells.get("desafioFinal")) == "sim"
    percentual = percentual_integralizacao(horas_vistas, config["horas_totais"], desafio_final)
    aluno_concluido = bool(desafio_final or percentual >= 100)
    meta_diaria = calcular_meta_diaria(
        horas_vistas,
        config["horas_totais"],
        config["prazo_final"],
        desafio_final=desafio_final,
        hoje=hoje,
    )

    decisao = texto(cells.get("decisao"))
    return {
        "nome": texto(cells.get("nome")) or email,
        "email": email,
        "emailNormalizado": email_normalizado,
        "alunoSlug": texto(cells.get("aluno")),
        "horasVistas": horas_vistas,
        "dataIngresso": formatar_data_br_date(data_entrada),
        "dataEntradaCurso": formatar_data_iso(data_entrada),
        "dataEntradaCursoFormatada": formatar_data_br_date(data_entrada),
        "decisao": decisao,
        "ativo": chave_campo(decisao) == "manter",
        "pdita": texto(cells.get("pdita")),
        "statusCruzamento": texto(cells.get("statusCruzamento")),
        "desafioFinal": desafio_final,
        "alunoConcluido": aluno_concluido,
        "horasTotaisCurso": config["horas_totais"],
        "percentualIntegralizacao": percentual,
        "certificados": build_certificados(
            cells,
            desafio_final,
            total_cursos_certificaveis=config["total_cursos_certificaveis"],
        ),
        "metaDiaria": meta_diaria,
    }


def status_curso_por_percentual(percentual, certificado_gerado=False):
    if certificado_gerado or percentual >= 100:
        return "Concluido"
    if percentual <= 0:
        return "Nao iniciado"
    return "Em andamento"


def montar_curso_neon(row):
    percentual = parse_percentual(row.get("percentual"))
    certificado_gerado = bool(row.get("certificado_gerado"))
    nome = texto(row.get("course_name"))
    course_id = texto(row.get("course_id")) or nome
    if not nome and not course_id:
        return None
    return {
        "curso": nome or course_id,
        "courseId": course_id,
        "status": texto(row.get("status")) or status_curso_por_percentual(percentual, certificado_gerado),
        "percentual": percentual,
        "certificadoGerado": certificado_gerado,
    }


def formatar_percentual_curso(valor):
    percentual = parse_percentual(valor)
    if float(percentual).is_integer():
        return f"{int(percentual)}%"
    return f"{percentual:.1f}%"


def texto_cursos_resumo(cursos):
    return "; ".join(
        f"{curso.get('curso')} ({formatar_percentual_curso(curso.get('percentual'))})"
        for curso in cursos
        if curso.get("curso")
    )


def build_certificados_neon(aluno_row, cursos, desafio_final=False, total_cursos_certificaveis=TOTAL_CURSOS_CERTIFICAVEIS):
    cursos = expandir_cursos_certificados(cursos, total_cursos_certificaveis)
    if desafio_final:
        grupos = agrupar_cursos(cursos)
        com_certificado = grupos["comCertificado"]
        return {
            "totalCursosCertificaveis": total_cursos_certificaveis,
            "cursosConcluidos": total_cursos_certificaveis,
            "certificadosGerados": total_cursos_certificaveis,
            "cursosEmAndamento": 0,
            "cursosNaoIniciados": 0,
            "cursosComCertificado": texto_cursos_resumo(com_certificado),
            "cursosSemCertificado": "",
            "cursos": cursos,
            "grupos": grupos,
        }

    grupos = agrupar_cursos(cursos)
    com_certificado = grupos["comCertificado"]
    sem_certificado = grupos["semCertificado"]
    return {
        "totalCursosCertificaveis": total_cursos_certificaveis,
        "cursosConcluidos": inteiro_seguro(aluno_row.get("cursos_concluidos"), len(grupos["concluidos"])),
        "certificadosGerados": inteiro_seguro(aluno_row.get("certificados_gerados"), len(com_certificado)),
        "cursosEmAndamento": inteiro_seguro(aluno_row.get("cursos_em_andamento"), len(grupos["emAndamento"])),
        "cursosNaoIniciados": inteiro_seguro(aluno_row.get("cursos_nao_iniciados"), len(grupos["naoIniciados"])),
        "cursosComCertificado": texto_cursos_resumo(com_certificado),
        "cursosSemCertificado": texto_cursos_resumo(sem_certificado),
        "cursos": cursos,
        "grupos": grupos,
    }


def montar_aluno_integralizacao_neon(aluno_row, cursos, config, hoje=None):
    email = texto(aluno_row.get("student_email"))
    email_normalizado = normalizar_email(email)
    if not email_normalizado:
        return None

    percentual = parse_percentual(aluno_row.get("consumo_percentual"))
    desafio_final = bool(aluno_row.get("desafio_final"))
    if desafio_final:
        percentual = 100

    horas_totais = config["horas_totais"]
    horas_vistas = round((horas_totais * percentual) / 100, 2)
    data_entrada = parse_data_excel(aluno_row.get("ingresso"))
    aluno_concluido = bool(desafio_final or percentual >= 100)
    meta_diaria = calcular_meta_diaria(
        horas_vistas,
        horas_totais,
        config["prazo_final"],
        desafio_final=desafio_final,
        hoje=hoje,
    )
    matricula = texto(aluno_row.get("matricula_pd"))

    return {
        "nome": texto(aluno_row.get("student_name")) or email,
        "email": email,
        "emailNormalizado": email_normalizado,
        "alunoSlug": email_normalizado.split("@", 1)[0],
        "horasVistas": horas_vistas,
        "dataIngresso": formatar_data_br_date(data_entrada),
        "dataEntradaCurso": formatar_data_iso(data_entrada),
        "dataEntradaCursoFormatada": formatar_data_br_date(data_entrada),
        "decisao": "",
        "ativo": True,
        "pdita": matricula,
        "matriculaPd": matricula,
        "cidade": texto(aluno_row.get("cidade")),
        "linkedStudentId": aluno_row.get("linked_student_id"),
        "statusCruzamento": "",
        "desafioFinal": desafio_final,
        "alunoConcluido": aluno_concluido,
        "horasTotaisCurso": horas_totais,
        "percentualIntegralizacao": percentual,
        "fonteConsumo": "neon",
        "certificados": build_certificados_neon(
            aluno_row,
            cursos,
            desafio_final,
            total_cursos_certificaveis=config["total_cursos_certificaveis"],
        ),
        "metaDiaria": meta_diaria,
    }


def montar_aluno_integralizacao_neon_resumo(aluno_row, config):
    email = texto(aluno_row.get("student_email"))
    email_normalizado = normalizar_email(email)
    if not email_normalizado:
        return None

    percentual = parse_percentual(aluno_row.get("consumo_percentual"))
    desafio_final = bool(aluno_row.get("desafio_final"))
    if desafio_final:
        percentual = 100

    data_entrada = parse_data_excel(aluno_row.get("ingresso"))
    matricula = texto(aluno_row.get("matricula_pd"))
    horas_totais = config["horas_totais"]
    horas_vistas = round((horas_totais * percentual) / 100, 2)

    return {
        "nome": texto(aluno_row.get("student_name")) or email,
        "email": email,
        "emailNormalizado": email_normalizado,
        "alunoSlug": email_normalizado.split("@", 1)[0],
        "horasVistas": horas_vistas,
        "dataIngresso": formatar_data_br_date(data_entrada),
        "dataEntradaCurso": formatar_data_iso(data_entrada),
        "dataEntradaCursoFormatada": formatar_data_br_date(data_entrada),
        "decisao": "",
        "ativo": True,
        "pdita": matricula,
        "matriculaPd": matricula,
        "cidade": texto(aluno_row.get("cidade")),
        "linkedStudentId": aluno_row.get("linked_student_id"),
        "statusCruzamento": "",
        "desafioFinal": desafio_final,
        "alunoConcluido": bool(desafio_final or percentual >= 100),
        "horasTotaisCurso": horas_totais,
        "percentualIntegralizacao": percentual,
        "fonteConsumo": "neon",
    }


def fonte_neon(run, alunos, config):
    finished_at = run.get("finished_at") or run.get("created_at")
    started_at = run.get("started_at")
    return {
        "tipo": "neon",
        "runId": run.get("id"),
        "status": run.get("status"),
        "startedAt": started_at.isoformat(timespec="seconds") if started_at else "",
        "finishedAt": finished_at.isoformat(timespec="seconds") if finished_at else "",
        "atualizadoEm": finished_at.isoformat(timespec="seconds") if finished_at else "",
        "horasTotaisCurso": config["horas_totais"],
        "prazoFinal": config["prazo_final"].isoformat(),
        "prazoFinalFormatado": formatar_data_br_date(config["prazo_final"]),
        "totalAlunos": len(alunos),
        "totalCursosCertificaveis": config["total_cursos_certificaveis"],
    }


def ler_neon_integralizacao(config, hoje=None):
    database_url = config.get("database_url")
    if not database_url:
        raise IntegralizacaoNeonIndisponivel(
            "DATABASE_URL nao configurada para leitura de consumo no Neon."
        )

    conn = None
    try:
        conn = psycopg2.connect(database_url)
        run = get_latest_successful_run(conn)
        if not run:
            raise IntegralizacaoNeonSemDados(
                "Nao ha dados de consumo no banco. Nenhuma execucao success foi encontrada."
            )

        alunos_rows = get_consumption_students_from_run(conn, run["id"])
        cursos_rows = get_consumption_courses_from_run(conn, run["id"])
    except IntegralizacaoNeonSemDados:
        raise
    except psycopg2.Error as exc:
        raise IntegralizacaoNeonIndisponivel(
            "Nao foi possivel ler os dados de consumo no Neon."
        ) from exc
    finally:
        if conn:
            conn.close()

    cursos_por_email = {}
    for row in cursos_rows:
        curso = montar_curso_neon(row)
        if not curso:
            continue
        email = normalizar_email(row.get("student_email"))
        if email:
            cursos_por_email.setdefault(email, []).append(curso)

    alunos = []
    for aluno_row in alunos_rows:
        email = normalizar_email(aluno_row.get("student_email"))
        aluno = montar_aluno_integralizacao_neon(
            aluno_row,
            cursos_por_email.get(email, []),
            config,
            hoje=hoje,
        )
        if aluno:
            alunos.append(aluno)

    return {
        "alunos": alunos,
        "porEmail": {aluno["emailNormalizado"]: aluno for aluno in alunos},
        "fonte": fonte_neon(run, alunos, config),
    }


def ler_neon_integralizacao_resumo(config):
    database_url = config.get("database_url")
    if not database_url:
        raise IntegralizacaoNeonIndisponivel(
            "DATABASE_URL nao configurada para leitura de consumo no Neon."
        )

    conn = None
    try:
        conn = psycopg2.connect(database_url)
        run = get_latest_successful_run(conn)
        if not run:
            raise IntegralizacaoNeonSemDados(
                "Nao ha dados de consumo no banco. Nenhuma execucao success foi encontrada."
            )

        alunos_rows = get_consumption_students_from_run(conn, run["id"])
    except IntegralizacaoNeonSemDados:
        raise
    except psycopg2.Error as exc:
        raise IntegralizacaoNeonIndisponivel(
            "Nao foi possivel ler os dados de consumo no Neon."
        ) from exc
    finally:
        if conn:
            conn.close()

    alunos = []
    for aluno_row in alunos_rows:
        aluno = montar_aluno_integralizacao_neon_resumo(aluno_row, config)
        if aluno:
            alunos.append(aluno)

    return {
        "alunos": alunos,
        "porEmail": {aluno["emailNormalizado"]: aluno for aluno in alunos},
        "fonte": fonte_neon(run, alunos, config),
    }


def ler_neon_integralizacao_aluno(config, email, hoje=None):
    database_url = config.get("database_url")
    if not database_url:
        raise IntegralizacaoNeonIndisponivel(
            "DATABASE_URL nao configurada para leitura de consumo no Neon."
        )

    conn = None
    try:
        conn = psycopg2.connect(database_url)
        run = get_latest_successful_run(conn)
        if not run:
            raise IntegralizacaoNeonSemDados(
                "Nao ha dados de consumo no banco. Nenhuma execucao success foi encontrada."
            )

        aluno_row = get_consumption_student_from_run_by_email(conn, run["id"], email)
        if not aluno_row:
            return {
                "aluno": None,
                "fonte": fonte_neon(run, [], config),
            }
        cursos_rows = get_consumption_courses_from_run_by_email(conn, run["id"], email)
    except IntegralizacaoNeonSemDados:
        raise
    except psycopg2.Error as exc:
        raise IntegralizacaoNeonIndisponivel(
            "Nao foi possivel ler os dados de consumo no Neon."
        ) from exc
    finally:
        if conn:
            conn.close()

    cursos = []
    for row in cursos_rows:
        curso = montar_curso_neon(row)
        if curso:
            cursos.append(curso)

    aluno = montar_aluno_integralizacao_neon(aluno_row, cursos, config, hoje=hoje)
    return {
        "aluno": aluno,
        "fonte": fonte_neon(run, [aluno] if aluno else [], config),
    }


def ler_planilha_integralizacao(config, hoje=None):
    xlsx_path = config["xlsx_path"]
    if not xlsx_path.is_file():
        raise IntegralizacaoArquivoNaoEncontrado(
            f"Planilha de consumo não encontrada: {config['xlsx_path_configurado']}"
        )

    try:
        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        raise IntegralizacaoPlanilhaInvalida("Não foi possível abrir a planilha de consumo.") from exc

    if config["sheet_name"] not in workbook.sheetnames:
        workbook.close()
        raise IntegralizacaoPlanilhaInvalida(
            f"Aba \"{config['sheet_name']}\" não encontrada na planilha de consumo."
        )

    sheet = workbook[config["sheet_name"]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        raise IntegralizacaoPlanilhaInvalida("A planilha de consumo está vazia.")

    headers = [texto(item) for item in rows[0]]
    indices = header_index(headers)
    colunas = {
        "email": localizar_coluna(indices, "Email"),
        "aluno": localizar_coluna(indices, "Aluno"),
        "horasVistas": localizar_coluna(indices, "Horas vistas"),
        "nome": localizar_coluna(indices, "Nome"),
        "dataEntrada": localizar_coluna(indices, "Data de entrada"),
        "decisao": localizar_coluna(indices, "Decisão", "Decisao"),
        "pdita": localizar_coluna(indices, "PDITA"),
        "statusCruzamento": localizar_coluna(indices, "Status do cruzamento"),
        "desafioFinal": localizar_coluna(indices, "Desafio Final"),
        "cursosConcluidos": localizar_coluna(indices, "Cursos concluídos", "Cursos concluidos"),
        "certificadosGerados": localizar_coluna(indices, "Certificados gerados"),
        "cursosEmAndamento": localizar_coluna(indices, "Cursos em andamento"),
        "cursosNaoIniciados": localizar_coluna(indices, "Cursos não iniciados", "Cursos nao iniciados"),
        "cursosComCertificado": localizar_coluna(indices, "Cursos com certificado"),
        "cursosSemCertificado": localizar_coluna(indices, "Cursos sem certificado"),
        "cursosDetalhesJson": localizar_coluna(indices, "Cursos detalhes JSON"),
    }
    if colunas["email"] is None:
        raise IntegralizacaoPlanilhaInvalida("Coluna Email não encontrada na aba de consumo.")

    alunos = []
    emails_vistos = set()
    duplicados = []
    for row in rows[1:]:
        cells = {campo: valor_linha(row, indice) for campo, indice in colunas.items()}
        aluno = montar_aluno_integralizacao(cells, config, hoje=hoje)
        if not aluno:
            continue
        email_normalizado = aluno["emailNormalizado"]
        if email_normalizado in emails_vistos:
            duplicados.append(email_normalizado)
        emails_vistos.add(email_normalizado)
        alunos.append(aluno)

    atualizado_em = datetime.now().isoformat(timespec="seconds")
    return {
        "alunos": alunos,
        "porEmail": {aluno["emailNormalizado"]: aluno for aluno in alunos},
        "fonte": {
            "tipo": "xlsx",
            "caminhoConfigurado": config["xlsx_path_configurado"],
            "aba": config["sheet_name"],
            "horasTotaisCurso": config["horas_totais"],
            "prazoFinal": config["prazo_final"].isoformat(),
            "prazoFinalFormatado": formatar_data_br_date(config["prazo_final"]),
            "totalLinhas": len(rows) - 1,
            "totalAlunos": len(alunos),
            "emailsDuplicados": sorted(set(duplicados)),
            "atualizadoEm": atualizado_em,
            "totalCursosCertificaveis": config["total_cursos_certificaveis"],
        },
    }


def carregar_por_fonte(config, hoje=None, incluir_cursos=True):
    modo = config["source_mode"]
    if modo == "xlsx":
        return ler_planilha_integralizacao(config, hoje=hoje)
    if modo == "neon":
        if incluir_cursos:
            return ler_neon_integralizacao(config, hoje=hoje)
        return ler_neon_integralizacao_resumo(config)

    try:
        dados = ler_neon_integralizacao(config, hoje=hoje) if incluir_cursos else ler_neon_integralizacao_resumo(config)
        dados["fonte"]["modoFonte"] = "auto"
        return dados
    except (IntegralizacaoNeonSemDados, IntegralizacaoNeonIndisponivel) as exc:
        dados = ler_planilha_integralizacao(config, hoje=hoje)
        dados["fonte"]["modoFonte"] = "auto"
        dados["fonte"]["fallback"] = "xlsx"
        dados["fonte"]["neonFallbackMotivo"] = exc.__class__.__name__
        return dados


def carregar_integralizacao(env=None, usar_cache=True, hoje=None, incluir_cursos=True, **overrides):
    config = config_integralizacao(env=env, **overrides)
    key = cache_key(config, hoje, incluir_cursos=incluir_cursos)
    agora = time.time()
    if usar_cache and _CACHE["key"] == key and _CACHE["data"] and agora < _CACHE["expires_at"]:
        return copy.deepcopy(_CACHE["data"])

    dados = carregar_por_fonte(config, hoje=hoje, incluir_cursos=incluir_cursos)
    if usar_cache and config["cache_ttl"] > 0:
        _CACHE["key"] = key
        _CACHE["expires_at"] = agora + config["cache_ttl"]
        _CACHE["data"] = copy.deepcopy(dados)
    return dados


def carregar_integralizacao_aluno(email, env=None, hoje=None, **overrides):
    config = config_integralizacao(env=env, **overrides)
    modo = config["source_mode"]
    if modo == "xlsx":
        dados = carregar_por_fonte(config, hoje=hoje, incluir_cursos=True)
        return {
            "aluno": buscar_por_email(dados, email),
            "fonte": dados.get("fonte"),
        }

    try:
        resultado = ler_neon_integralizacao_aluno(config, email, hoje=hoje)
        if resultado.get("fonte") is not None and modo == "auto":
            resultado["fonte"]["modoFonte"] = "auto"
        return resultado
    except (IntegralizacaoNeonSemDados, IntegralizacaoNeonIndisponivel) as exc:
        if modo == "neon":
            raise
        dados = ler_planilha_integralizacao(config, hoje=hoje)
        dados["fonte"]["modoFonte"] = "auto"
        dados["fonte"]["fallback"] = "xlsx"
        dados["fonte"]["neonFallbackMotivo"] = exc.__class__.__name__
        return {
            "aluno": buscar_por_email(dados, email),
            "fonte": dados.get("fonte"),
        }


def aluno_pd_resumo(aluno):
    if not aluno:
        return None
    return {
        "id": aluno.get("id"),
        "matricula": aluno.get("matricula") or "",
        "nome": aluno.get("nome") or "",
        "email": aluno.get("email") or "",
        "monitor": aluno.get("monitor") or "",
        "status": aluno.get("status") or "",
    }


def cruzar_com_alunos_pd(alunos_integralizacao, alunos_pd):
    pd_por_email = {}
    pd_por_id = {}
    duplicados_pd = set()
    for aluno in alunos_pd:
        aluno_id = aluno.get("id")
        if aluno_id is not None:
            pd_por_id[str(aluno_id)] = aluno
        email = normalizar_email(aluno.get("email"))
        if not email:
            continue
        if email in pd_por_email:
            duplicados_pd.add(email)
            continue
        pd_por_email[email] = aluno

    resultado = []
    vinculados = 0
    nao_vinculados = 0
    for aluno in alunos_integralizacao:
        item = copy.deepcopy(aluno)
        linked_student_id = item.get("linkedStudentId")
        aluno_pd = pd_por_id.get(str(linked_student_id)) if linked_student_id is not None else None
        if not aluno_pd:
            aluno_pd = pd_por_email.get(item["emailNormalizado"])
        item["vinculado"] = bool(aluno_pd)
        item["alunoPd"] = aluno_pd_resumo(aluno_pd)
        if aluno_pd:
            if item.get("fonteConsumo") == "neon" and aluno_pd.get("status"):
                item["decisao"] = texto(aluno_pd.get("status"))
                item["ativo"] = chave_campo(aluno_pd.get("status")) == "manter"
            if not item.get("pdita"):
                item["pdita"] = texto(aluno_pd.get("matricula"))
            vinculados += 1
        else:
            nao_vinculados += 1
        resultado.append(item)

    return {
        "alunos": resultado,
        "resumoVinculos": {
            "vinculados": vinculados,
            "naoVinculados": nao_vinculados,
            "emailsDuplicadosPdReports": sorted(duplicados_pd),
        },
    }


def buscar_por_email(dados_integralizacao, email):
    email_normalizado = normalizar_email(email)
    if not email_normalizado:
        return None
    aluno = dados_integralizacao.get("porEmail", {}).get(email_normalizado)
    return copy.deepcopy(aluno) if aluno else None


def montar_resumo_geral(alunos):
    total = len(alunos)
    vinculados = len([item for item in alunos if item.get("vinculado")])
    nao_vinculados = total - vinculados
    concluidos = len([item for item in alunos if item.get("alunoConcluido")])
    ativos = len([item for item in alunos if item.get("ativo")])
    return {
        "total": total,
        "ativos": ativos,
        "inativos": total - ativos,
        "vinculados": vinculados,
        "naoVinculados": nao_vinculados,
        "concluidos": concluidos,
    }
