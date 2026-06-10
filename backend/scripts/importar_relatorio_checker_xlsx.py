import argparse
import json
import os
import subprocess
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


BACKEND_DIR = Path(__file__).resolve().parents[1]
PROJECT_ROOT = BACKEND_DIR.parent
VENV_PYTHON = BACKEND_DIR / "venv" / "Scripts" / "python.exe"
if VENV_PYTHON.exists() and Path(sys.executable).resolve() != VENV_PYTHON.resolve():
    raise SystemExit(subprocess.call([str(VENV_PYTHON), *sys.argv]))

if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

import psycopg2
from dotenv import load_dotenv

from course_checker import (
    CourseCheckerError,
    build_source_files_info,
    course_name_is_excluded_from_consumption,
    create_consumption_run,
    load_existing_consumption_enrichment,
    mark_consumption_run_error,
    normalize_email,
    parse_percent_input,
    parse_total_certifiable,
    persist_consumption_run,
    text,
)
from course_rules import COURSE_CONSUMPTION_TOTAL_CERTIFIABLE


load_dotenv(dotenv_path=BACKEND_DIR / ".env", override=True)

CONFIRM_FLAG = "--confirmar-importacao"
DEFAULT_REPORT_PATH = "checker/relatorio_final.xlsx"


def conectar_db():
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        raise CourseCheckerError("DATABASE_URL nao configurada no backend/.env.")
    return psycopg2.connect(database_url)


def resolve_project_path(value):
    path = Path(value or "")
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path


def parse_args():
    parser = argparse.ArgumentParser(
        description="Importa o relatorio_final.xlsx do checker para o Neon."
    )
    parser.add_argument(
        CONFIRM_FLAG,
        action="store_true",
        dest="confirmar_importacao",
        help="Confirma que o arquivo importado e um relatorio temporario aprovado para carga no Neon.",
    )
    return parser.parse_args()


def load_report_path():
    report_path = os.getenv("CHECKER_REPORT_XLSX_PATH", DEFAULT_REPORT_PATH)
    return resolve_project_path(report_path)


def ensure_report_exists(report_path):
    if not report_path.is_file():
        raise CourseCheckerError(f"Relatorio do checker nao encontrado: {report_path}")


def header_index(headers):
    return {text(header): index for index, header in enumerate(headers)}


def pick_value(row, indices, *candidates):
    for candidate in candidates:
        index = indices.get(text(candidate))
        if index is not None and index < len(row):
            value = row[index]
            if value is not None and value != "":
                return value
    return None


def load_report_rows(report_path):
    workbook = load_workbook(report_path, read_only=True, data_only=True)
    try:
        resumo_name = "Resumo por aluno"
        cursos_name = "Cursos por aluno"
        if resumo_name not in workbook.sheetnames or cursos_name not in workbook.sheetnames:
            raise CourseCheckerError(
                "O relatorio precisa conter as abas 'Resumo por aluno' e 'Cursos por aluno'."
            )

        resumo_ws = workbook[resumo_name]
        cursos_ws = workbook[cursos_name]

        resumo_rows = list(resumo_ws.iter_rows(values_only=True))
        cursos_rows = list(cursos_ws.iter_rows(values_only=True))
        if len(resumo_rows) < 2 or len(cursos_rows) < 2:
            raise CourseCheckerError("O relatorio do checker esta vazio ou incompleto.")

        resumo_headers = header_index(resumo_rows[0])
        cursos_headers = header_index(cursos_rows[0])

        resumo = {}
        resumo_duplicados = 0
        for row in resumo_rows[1:]:
            email = normalize_email(pick_value(row, resumo_headers, "Email"))
            if not email:
                continue
            if email in resumo:
                resumo_duplicados += 1
                continue
            resumo[email] = {
                "email": email,
                "nome": text(pick_value(row, resumo_headers, "Aluno")) or email,
            }

        cursos_por_email = defaultdict(list)
        catalog_by_course = {}
        cursos_ignorados = 0
        cursos_invalidos = 0
        for ordem, row in enumerate(cursos_rows[1:], start=1):
            email = normalize_email(pick_value(row, cursos_headers, "Email"))
            if not email:
                cursos_invalidos += 1
                continue

            course_id = text(pick_value(row, cursos_headers, "Course ID", "CourseId", "course_id"))
            course_name = text(pick_value(row, cursos_headers, "Curso", "Course", "course_name"))
            if not course_id:
                course_id = course_name
            if not course_name:
                course_name = course_id
            if not course_name:
                cursos_invalidos += 1
                continue
            if course_name_is_excluded_from_consumption(course_name) or course_name_is_excluded_from_consumption(course_id):
                cursos_ignorados += 1
                catalog_by_course.setdefault(course_id, {
                    "courseId": course_id,
                    "courseName": course_name,
                    "certificavel": False,
                    "ignored": True,
                    "ordem": ordem,
                })
                continue

            percentual_100, _ = parse_percent_input(pick_value(row, cursos_headers, "Percentual"))
            status = text(pick_value(row, cursos_headers, "Status")) or (
                "Concluído" if percentual_100 >= 100 else "Em andamento" if percentual_100 > 0 else "Não iniciado"
            )
            certificado_gerado = bool(pick_value(row, cursos_headers, "Certificado gerado"))

            catalog_by_course.setdefault(course_id, {
                "courseId": course_id,
                "courseName": course_name,
                "certificavel": True,
                "ignored": False,
                "ordem": ordem,
            })
            cursos_por_email[email].append({
                "courseId": course_id,
                "courseName": course_name,
                "status": status,
                "percentual": percentual_100,
                "certificadoGerado": certificado_gerado,
            })

        total_linhas_cursos = len(cursos_rows) - 1
        return {
            "resumo": resumo,
            "cursos_por_email": cursos_por_email,
            "catalogo": list(catalog_by_course.values()),
            "warnings": {
                "resumo_duplicados": resumo_duplicados,
                "cursos_ignorados": cursos_ignorados,
                "cursos_invalidos": cursos_invalidos,
                "total_linhas_cursos": total_linhas_cursos,
            },
        }
    finally:
        workbook.close()


def normalize_student_courses(student, total_certifiable):
    cursos = student.get("cursos", [])
    concluidos = sum(1 for curso in cursos if text(curso.get("status")).lower().startswith("conclu"))
    em_andamento = sum(1 for curso in cursos if text(curso.get("status")) == "Em andamento")
    nao_iniciados = sum(1 for curso in cursos if text(curso.get("status")) == "Não iniciado")
    certificados = sum(1 for curso in cursos if curso.get("certificadoGerado"))
    total_contado = concluidos + em_andamento + nao_iniciados
    if total_contado < total_certifiable:
        nao_iniciados += total_certifiable - total_contado
    elif total_contado > total_certifiable:
        excesso = total_contado - total_certifiable
        nao_iniciados = max(0, nao_iniciados - excesso)

    consumo = round(sum(float(curso.get("percentual") or 0) for curso in cursos) / total_certifiable, 2)
    if student.get("desafioFinal"):
        consumo = 100
        concluidos = total_certifiable
        em_andamento = 0
        nao_iniciados = 0
        certificados = total_certifiable
        for curso in cursos:
            curso["status"] = "Concluído"
            curso["percentual"] = 100
            curso["certificadoGerado"] = True

    student["consumoPercentual"] = consumo
    student["certificadosGerados"] = min(certificados, total_certifiable)
    student["cursosConcluidos"] = min(concluidos, total_certifiable)
    student["cursosEmAndamento"] = min(em_andamento, total_certifiable)
    student["cursosNaoIniciados"] = min(nao_iniciados, total_certifiable)
    student["cursosSemCertificado"] = max(total_certifiable - student["certificadosGerados"], 0)
    return student


def build_payload_from_report(report_path, enrichment_by_email, total_certifiable):
    report = load_report_rows(report_path)
    students = []
    emails = set(report["resumo"].keys()) | set(report["cursos_por_email"].keys())
    for email in sorted(emails):
        base = report["resumo"].get(email, {"email": email, "nome": email})
        student = {
            "email": email,
            "nome": base.get("nome") or email,
            "desafioFinal": False,
            "ingresso": None,
            "cursos": report["cursos_por_email"].get(email, []),
        }
        enrich = enrichment_by_email.get(email) or {}
        student["desafioFinal"] = bool(enrich.get("desafioFinal", False))
        student["ingresso"] = enrich.get("ingresso")
        normalize_student_courses(student, total_certifiable)
        students.append(student)

    payload = {
        "students": students,
        "warnings": [],
        "courseCatalog": report["catalogo"],
        "sourceFilesInfo": build_source_files_info({"checker_report": report_path}),
        "totals": {
            "students": len(students),
            "courses": sum(len(student.get("cursos", [])) for student in students),
            "resumoDuplicados": report["warnings"]["resumo_duplicados"],
            "cursosIgnorados": report["warnings"]["cursos_ignorados"],
            "cursosInvalidos": report["warnings"]["cursos_invalidos"],
            "linhasCursos": report["warnings"]["total_linhas_cursos"],
        },
        "source": "checker_relatorio_xlsx",
    }

    if report["warnings"]["resumo_duplicados"]:
        payload["warnings"].append(
            f"{report['warnings']['resumo_duplicados']} e-mails duplicados foram ignorados no resumo do relatorio."
        )
    if report["warnings"]["cursos_ignorados"]:
        payload["warnings"].append(
            f"{report['warnings']['cursos_ignorados']} linhas de cursos foram ignoradas por regra oficial."
        )
    if report["warnings"]["cursos_invalidos"]:
        payload["warnings"].append(
            f"{report['warnings']['cursos_invalidos']} linhas de cursos invalidas foram descartadas."
        )

    return payload


def print_summary(report_path, payload, enrichment_available, total_certifiable):
    print("Importacao temporaria do relatorio do checker")
    print(f"Arquivo: {report_path.name}")
    print(f"Total oficial de cursos certificaveis: {total_certifiable}")
    print(f"Alunos preparados: {payload['totals']['students']}")
    print(f"Cursos preparados: {payload['totals']['courses']}")
    print(f"Enriquecimento XLSX atual: {'disponivel' if enrichment_available else 'indisponivel'}")
    print("Os e-mails individuais nao sao exibidos neste resumo.")


def main():
    args = parse_args()
    total_certifiable = parse_total_certifiable(
        os.getenv("COURSE_CONSUMPTION_TOTAL_CERTIFIABLE", str(COURSE_CONSUMPTION_TOTAL_CERTIFIABLE))
    )
    report_path = load_report_path()
    ensure_report_exists(report_path)

    if not args.confirmar_importacao:
        print("Importacao abortada por seguranca.")
        print(f"Para executar, rode novamente com {CONFIRM_FLAG}.")
        raise SystemExit(1)

    conn = conectar_db()
    run_id = None
    payload = None
    enrichment_available = False
    warnings = []
    try:
        run_id = create_consumption_run(
            conn,
            source_files_info=build_source_files_info({"checker_report": report_path}),
        )

        try:
            enrichment_by_email = load_existing_consumption_enrichment(env=os.environ)
            enrichment_available = True
        except Exception:
            enrichment_by_email = {}

        payload = build_payload_from_report(report_path, enrichment_by_email, total_certifiable)
        warnings = list(payload.get("warnings", []))
        print_summary(report_path, payload, enrichment_available, total_certifiable)

        result = persist_consumption_run(
            conn,
            payload,
            run_id=run_id,
            source_files_info=payload.get("sourceFilesInfo"),
        )
        print(json.dumps({
            "status": result["status"],
            "run_id": result["run_id"],
            "total_alunos": result["students"],
            "total_cursos": result["courses"],
            "warnings": result["warnings"],
        }, ensure_ascii=False, indent=2))
    except Exception as exc:
        if run_id is not None:
            warnings_to_save = warnings or (payload.get("warnings", []) if payload else [])
            try:
                mark_consumption_run_error(conn, run_id, str(exc), warnings_to_save)
            except Exception:
                conn.rollback()
        print(json.dumps({
            "status": "error",
            "run_id": run_id,
            "erro": str(exc),
            "warnings": warnings,
        }, ensure_ascii=False, indent=2))
        raise SystemExit(1) from exc
    finally:
        conn.close()


if __name__ == "__main__":
    main()
