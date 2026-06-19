from contextlib import contextmanager
import json
import os
from pathlib import Path
import shutil
import subprocess
import sys
import tempfile

from openpyxl import Workbook


BACKEND_DIR = Path(__file__).resolve().parents[1]
VENV_PYTHON = BACKEND_DIR / 'venv' / 'Scripts' / 'python.exe'
if VENV_PYTHON.exists() and Path(sys.executable).resolve() != VENV_PYTHON.resolve():
    raise SystemExit(subprocess.call([str(VENV_PYTHON), *sys.argv]))

if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

import app as app_module
from checker_report_importer import build_payload_from_report
from consumption_update_service import (
    SOURCE_TYPE_FULL_CHECKER_LOCAL_CSV,
    build_payload_from_checker_paths,
    validate_checker_inputs,
)


ADMIN = {'id': 1, 'nome': 'Admin', 'email': 'admin@example.com', 'role': 'admin', 'ativo': True}
MONITOR = {'id': 2, 'nome': 'Monitor', 'email': 'monitor@example.com', 'role': 'monitor', 'ativo': True}
PSICOLOGA = {'id': 3, 'nome': 'Psicóloga', 'email': 'psicologa@example.com', 'role': 'psicologa', 'ativo': True}
PREFEITURA = {'id': 4, 'nome': 'Itabira - Prefeitura', 'email': 'prefeitura@example.com', 'role': 'prefeitura_itabira', 'ativo': True}


@contextmanager
def patched(module, **replacements):
    original = {}
    for name, value in replacements.items():
        original[name] = getattr(module, name)
        setattr(module, name, value)
    try:
        yield
    finally:
        for name, value in original.items():
            setattr(module, name, value)


def assert_equal(descricao, recebido, esperado):
    if recebido != esperado:
        raise AssertionError(f'{descricao}: esperado {esperado!r}, recebido {recebido!r}')
    print(f'OK - {descricao}')


def assert_true(descricao, valor):
    if not valor:
        raise AssertionError(f'{descricao}: esperado True, recebido {valor!r}')
    print(f'OK - {descricao}')


def create_report_xlsx(path, *, include_intensivao=True, username_like=False):
    workbook = Workbook()
    resumo = workbook.active
    resumo.title = 'Resumo por aluno'
    resumo.append(['Aluno', 'Email', 'Cursos concluídos', 'Cursos em andamento', 'Cursos não iniciados', 'Certificados gerados'])
    resumo.append([
        'alisson.goncalves' if username_like else 'Alisson Vinicius Ferreira Goncalves',
        'alisson.goncalves@example.com',
        1,
        1,
        1,
        1,
    ])

    cursos = workbook.create_sheet('Cursos por aluno')
    cursos.append(['Aluno', 'Email', 'Curso', 'Course ID', 'Status', 'Percentual', 'Certificado gerado'])
    cursos.append(['alisson.goncalves', 'alisson.goncalves@example.com', 'Python 1', 'course-python-1', 'Em andamento', 0.14, 'Não'])
    cursos.append(['alisson.goncalves', 'alisson.goncalves@example.com', 'Python 2', 'course-python-2', 'Concluído', 0.83, 'Sim'])
    if include_intensivao:
        cursos.append(['alisson.goncalves', 'alisson.goncalves@example.com', 'Intensivão Desenvolve 2025', 'course-intensivao', 'Concluído', 1, 'Sim'])
    workbook.save(path)


def create_minimal_report(path):
    workbook = Workbook()
    resumo = workbook.active
    resumo.title = 'Resumo por aluno'
    resumo.append(['Aluno', 'Email', 'Cursos concluídos', 'Cursos em andamento', 'Cursos não iniciados', 'Certificados gerados'])
    resumo.append(['Aluno Teste', 'teste@example.com', 1, 0, 0, 0])
    cursos = workbook.create_sheet('Cursos por aluno')
    cursos.append(['Aluno', 'Email', 'Curso', 'Course ID', 'Status', 'Percentual', 'Certificado gerado'])
    cursos.append(['Aluno Teste', 'teste@example.com', 'Python 1', 'course-python-1', 'Concluído', 1, 'Sim'])
    workbook.save(path)


def create_invalid_report(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Outra aba'
    workbook.save(path)


def create_checker_files(tmpdir):
    users = tmpdir / 'users.json'
    catalog = tmpdir / 'cursos_new.json'
    ignore = tmpdir / 'ignore_courses.json'
    grades = tmpdir / 'all_grades.json'
    certificates = tmpdir / 'certificados_20260602_141114.csv'
    users.write_text(json.dumps({
        'alisson': {'email': 'alisson.goncalves@example.com', 'name': 'alisson.goncalves'},
    }), encoding='utf-8')
    catalogo = {
        'course-v1:test+python1+2026': {'course_name': 'Python 1', 'certificavel': True, 'ordem': 1},
    }
    for indice in range(2, 23):
        catalogo[f'course-v1:test+oficial{indice:02d}+2026'] = {
            'course_name': f'Curso Oficial {indice:02d}',
            'certificavel': True,
            'ordem': indice,
        }
    catalogo['course-v1:test+intensivao+2026'] = {
        'course_name': 'Intensivao Desenvolve 2025',
        'certificavel': True,
        'ordem': 23,
    }
    catalog.write_text(json.dumps(catalogo), encoding='utf-8')
    ignore.write_text('[]', encoding='utf-8')
    grades.write_text(json.dumps({
        'course-v1:test+python1+2026': [{'username': 'alisson', 'percent': 1}],
        'course-v1:test+intensivao+2026': [{'username': 'alisson', 'percent': 1}],
    }), encoding='utf-8')
    certificates.write_text(
        'username,course_id,status,is_passing\n'
        'alisson,course-v1:test+python1+2026,downloadable,true\n'
        'alisson,course-v1:test+intensivao+2026,downloadable,true\n',
        encoding='utf-8',
    )
    return {
        'users': users,
        'catalog': catalog,
        'ignore': ignore,
        'grades': grades,
        'certificates': certificates,
    }


def make_temp_wrapper():
    file_obj = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')

    class Wrapper:
        name = file_obj.name

        def close(self):
            file_obj.close()

    return Wrapper()


def post_upload(client, report_path):
    with report_path.open('rb') as file_obj:
        return client.post(
            '/api/admin/consumo/importar-relatorio',
            data={'arquivo': (file_obj, report_path.name)},
            content_type='multipart/form-data',
        )


def post_checker_upload(client, grades_path, certificates_path):
    data = {}
    handles = []
    try:
        if grades_path:
            handles.append(grades_path.open('rb'))
            data['all_grades'] = (handles[-1], grades_path.name)
        if certificates_path:
            handles.append(certificates_path.open('rb'))
            data['certificates'] = (handles[-1], certificates_path.name)
        return client.post('/api/admin/consumo/atualizar', data=data, content_type='multipart/form-data')
    finally:
        for handle in handles:
            handle.close()


def test_payload_service(report_path):
    payload = build_payload_from_report(
        report_path,
        enrichment_by_email={
            'alisson.goncalves@example.com': {
                'nome': 'Alisson Vinicius Ferreira Goncalves',
                'desafioFinal': False,
                'ingresso': '2025-06-03',
            }
        },
        total_certifiable=22,
        original_filename='relatorio_final.xlsx',
    )
    aluno = payload['students'][0]
    assert_equal('origem do upload temporario mantem nome original', payload['sourceFilesInfo']['arquivoOriginal'], 'relatorio_final.xlsx')
    assert_equal('origem identifica checker report', payload['sourceType'], 'checker_report_xlsx')
    assert_equal('total certificavel oficial 22', payload['totalCertifiable'], 22)
    assert_equal('quantidade de cursos oficiais expandida', payload['sourceFilesInfo']['quantidadeCursos'], 22)
    assert_equal('xlsx expande aluno para 22 cursos', len(aluno['cursos']), 22)
    assert_true('curso intensivao ignorado no payload', all(curso['courseName'] != 'Intensivão Desenvolve 2025' for curso in aluno['cursos']))
    assert_equal('não iniciados ajusta para total oficial', aluno['cursosNaoIniciados'], 20)
    assert_equal('xlsx sem certificado bate com contador', sum(1 for curso in aluno['cursos'] if not curso['certificadoGerado']), aluno['cursosSemCertificado'])
    assert_equal('nome real do PD priorizado', aluno['nome'], 'Alisson Vinicius Ferreira Goncalves')


def test_checker_payload_service(paths):
    validate_checker_inputs(paths, max_grades_mb=150, max_certificates_mb=20)
    payload = build_payload_from_checker_paths(
        paths,
        total_certifiable=22,
        env={
            'INTEGRALIZACAO_XLSX_PATH': str(BACKEND_DIR / 'arquivo_inexistente.xlsx'),
            'INTEGRALIZACAO_SHEET_NAME': 'Resultado',
            'INTEGRALIZACAO_HORAS_TOTAIS': '154',
            'INTEGRALIZACAO_PRAZO_FINAL': '2026-11-30',
        },
    )
    aluno = payload['students'][0]
    assert_equal('checker manual source type', payload['sourceType'], SOURCE_TYPE_FULL_CHECKER_LOCAL_CSV)
    assert_equal('checker manual total 22', payload['totalCertifiable'], 22)
    assert_equal('checker manual expande 22 cursos e ignora intensivao', len(aluno['cursos']), 22)
    assert_equal('checker manual certificado valido', aluno['certificadosGerados'], 1)
    assert_equal('checker manual data do csv inferida', payload['sourceFilesInfo']['certificates']['csv_date'], '2026-06-02')
    assert_true(
        'checker manual warning csv antigo',
        any('02/06/2026' in warning for warning in payload['warnings']),
    )


def test_upload_success(client, report_path):
    created = {}

    def fake_importar(conn, path, **kwargs):
        created['path'] = Path(path)
        return {'status': 'success', 'run_id': 77, 'students': 1, 'courses': 2, 'warnings': ['ok']}

    temp_wrapper = make_temp_wrapper()
    with patched(app_module,
                 require_operational_admin=lambda: (ADMIN, None),
                 conectar_db=lambda: type('Conn', (), {'close': lambda self: None})(),
                 upload_consumo_bloqueado=lambda *chaves: False,
                 registrar_tentativa_upload_consumo=lambda *chaves: None,
                 consumo_upload_lock=lambda conn: True,
                 get_latest_active_run=lambda conn: None,
                 importar_relatorio_checker_xlsx=fake_importar,
                 consumo_upload_unlock=lambda conn: None,
                 tempfile=type('TempModule', (), {'NamedTemporaryFile': staticmethod(lambda *args, **kwargs: temp_wrapper)})):
        response = post_upload(client, report_path)

    data = response.get_json()
    assert_equal('upload admin retorna 200', response.status_code, 200)
    assert_equal('upload admin status success', data['status'], 'success')
    assert_equal('upload admin quantidade alunos', data['quantidade_alunos'], 1)
    assert_equal('upload admin quantidade cursos', data['quantidade_cursos'], 2)
    assert_equal('upload temporario enviado ao servico', created['path'], Path(temp_wrapper.name))
    assert_true('upload temporario removido', not Path(temp_wrapper.name).exists())


def test_checker_upload_start(client, paths):
    seen = {}

    def fake_stage(conn, grades_path, certificates_path, grades_name, certificates_name, **kwargs):
        seen['grades_exists'] = Path(grades_path).is_file()
        seen['certificates_exists'] = Path(certificates_path).is_file()
        Path(grades_path).unlink()
        Path(certificates_path).unlink()
        return {'run_id': 123, 'status': 'pending', 'message': 'Atualizacao recebida.'}

    with patched(app_module,
                 require_operational_admin=lambda: (ADMIN, None),
                 conectar_db=lambda: type('Conn', (), {'close': lambda self: None})(),
                 upload_consumo_bloqueado=lambda *chaves: False,
                 registrar_tentativa_upload_consumo=lambda *chaves: None,
                 limpar_tentativas_upload_consumo=lambda *chaves: None,
                 CONSUMPTION_PROCESSING_MODE='external',
                 stage_manual_consumption_update=fake_stage):
        response = post_checker_upload(client, paths['grades'], paths['certificates'])
    data = response.get_json()
    assert_equal('checker upload admin retorna 202', response.status_code, 202)
    assert_equal('checker upload cria pending', data['status'], 'pending')
    assert_equal('checker upload run id', data['run_id'], 123)
    assert_true('checker upload recebeu json temporario', seen['grades_exists'])
    assert_true('checker upload recebeu csv temporario', seen['certificates_exists'])


def test_checker_upload_sync_success(client, paths):
    seen = {}

    def fake_stage(conn, grades_path, certificates_path, grades_name, certificates_name, **kwargs):
        seen['grades_exists'] = Path(grades_path).is_file()
        seen['certificates_exists'] = Path(certificates_path).is_file()
        seen['grades_name'] = grades_name
        seen['certificates_name'] = certificates_name
        Path(grades_path).unlink()
        Path(certificates_path).unlink()
        return {'run_id': 124, 'status': 'pending', 'message': 'Atualizacao recebida.'}

    def fake_process(conn, run_id, **kwargs):
        seen['processed_run_id'] = run_id
        return {'run_id': run_id, 'status': 'success', 'students': 800, 'courses': 6230, 'warnings': ['csv antigo']}

    with patched(app_module,
                 require_operational_admin=lambda: (ADMIN, None),
                 conectar_db=lambda: type('Conn', (), {'close': lambda self: None})(),
                 upload_consumo_bloqueado=lambda *chaves: False,
                 registrar_tentativa_upload_consumo=lambda *chaves: None,
                 limpar_tentativas_upload_consumo=lambda *chaves: None,
                 CONSUMPTION_PROCESSING_MODE='sync',
                 stage_manual_consumption_update=fake_stage,
                 process_consumption_update_run=fake_process):
        response = post_checker_upload(client, paths['grades'], paths['certificates'])

    data = response.get_json()
    assert_equal('checker sync retorna 200', response.status_code, 200)
    assert_equal('checker sync status success', data['status'], 'success')
    assert_equal('checker sync run id', data['run_id'], 124)
    assert_equal('checker sync alunos', data['students'], 800)
    assert_equal('checker sync cursos', data['courses'], 6230)
    assert_equal('checker sync processou run criada', seen['processed_run_id'], 124)
    assert_true('checker sync recebeu json temporario', seen['grades_exists'])
    assert_true('checker sync recebeu csv temporario', seen['certificates_exists'])


def test_checker_upload_sync_error(client, paths):
    def fake_stage(conn, grades_path, certificates_path, grades_name, certificates_name, **kwargs):
        Path(grades_path).unlink()
        Path(certificates_path).unlink()
        return {'run_id': 125, 'status': 'pending', 'message': 'Atualizacao recebida.'}

    def fake_process(conn, run_id, **kwargs):
        raise TimeoutError(r'timeout lendo C:\segredo\certificados.csv')

    with patched(app_module,
                 require_operational_admin=lambda: (ADMIN, None),
                 conectar_db=lambda: type('Conn', (), {'close': lambda self: None})(),
                 upload_consumo_bloqueado=lambda *chaves: False,
                 registrar_tentativa_upload_consumo=lambda *chaves: None,
                 CONSUMPTION_PROCESSING_MODE='sync',
                 stage_manual_consumption_update=fake_stage,
                 process_consumption_update_run=fake_process):
        with patched(app_module.app.logger, exception=lambda *args, **kwargs: None):
            response = post_checker_upload(client, paths['grades'], paths['certificates'])

    data = response.get_json()
    assert_equal('checker sync erro retorna 500', response.status_code, 500)
    assert_equal('checker sync erro status', data['status'], 'error')
    assert_equal('checker sync erro run id', data['run_id'], 125)
    assert_true('checker sync timeout nao vira sucesso', data['status'] != 'success')
    assert_true('checker sync erro sanitizado', 'C:\\' not in data['erro'])


def test_upload_restrictions(client, report_path):
    for role in (MONITOR, PSICOLOGA, PREFEITURA):
        with patched(app_module, require_operational_admin=lambda role=role: (None, (app_module.jsonify({'erro': 'Apenas administradores podem executar esta ação.'}), 403))):
            response = post_upload(client, report_path)
            assert_equal(f'upload bloqueado para {role["role"]}', response.status_code, 403)

    with patched(app_module,
                 require_operational_admin=lambda: (ADMIN, None),
                 upload_consumo_bloqueado=lambda *chaves: False,
                 registrar_tentativa_upload_consumo=lambda *chaves: None):
        app_module.CONSUMO_UPLOAD_ATTEMPTS.clear()
        txt_path = BACKEND_DIR / '.upload_fake.txt'
        txt_path.write_text('nao xlsx', encoding='utf-8')
        try:
            with txt_path.open('rb') as file_obj:
                response = client.post('/api/admin/consumo/importar-relatorio', data={'arquivo': (file_obj, txt_path.name)}, content_type='multipart/form-data')
            assert_equal('upload nao xlsx recusado', response.status_code, 400)
        finally:
            if txt_path.exists():
                txt_path.unlink()

        invalid_report = BACKEND_DIR / '.upload_invalid.xlsx'
        create_invalid_report(invalid_report)
        try:
            response = post_upload(client, invalid_report)
            assert_equal('upload sem abas obrigatorias recusado', response.status_code, 400)
        finally:
            if invalid_report.exists():
                invalid_report.unlink()

        with patched(app_module, upload_consumo_bloqueado=lambda *chaves: False, registrar_tentativa_upload_consumo=lambda *chaves: None, consumo_upload_lock=lambda conn: True, get_latest_active_run=lambda conn: {'id': 1, 'status': 'running'}):
            response = post_upload(client, report_path)
            assert_equal('upload concorrente recusado', response.status_code, 409)


def test_checker_upload_restrictions(client, paths):
    for role in (MONITOR, PSICOLOGA, PREFEITURA):
        with patched(app_module, require_operational_admin=lambda role=role: (None, (app_module.jsonify({'erro': 'Apenas administradores podem executar esta acao.'}), 403))):
            response = post_checker_upload(client, paths['grades'], paths['certificates'])
            assert_equal(f'checker upload bloqueado para {role["role"]}', response.status_code, 403)

    with patched(app_module,
                 require_operational_admin=lambda: (ADMIN, None),
                 upload_consumo_bloqueado=lambda *chaves: False,
                 registrar_tentativa_upload_consumo=lambda *chaves: None):
        response = post_checker_upload(client, None, paths['certificates'])
        assert_equal('checker upload json ausente', response.status_code, 400)
        response = post_checker_upload(client, paths['grades'], None)
        assert_equal('checker upload csv ausente', response.status_code, 400)

        fake_txt = paths['grades'].with_suffix('.txt')
        fake_txt.write_text('{}', encoding='utf-8')
        try:
            response = post_checker_upload(client, fake_txt, paths['certificates'])
            assert_equal('checker upload extensao invalida', response.status_code, 400)
        finally:
            fake_txt.unlink(missing_ok=True)

        with patched(app_module, CONSUMPTION_UPLOAD_MAX_GRADES_MB=0):
            response = post_checker_upload(client, paths['grades'], paths['certificates'])
            assert_equal('checker upload arquivo excede limite', response.status_code, 413)

    def fake_conflict(*args, **kwargs):
        raise app_module.ConsumptionUpdateConflictError('Ja existe uma atualizacao de consumo em andamento.')

    with patched(app_module,
                 require_operational_admin=lambda: (ADMIN, None),
                 conectar_db=lambda: type('Conn', (), {'close': lambda self: None})(),
                 upload_consumo_bloqueado=lambda *chaves: False,
                 registrar_tentativa_upload_consumo=lambda *chaves: None,
                 stage_manual_consumption_update=fake_conflict):
        response = post_checker_upload(client, paths['grades'], paths['certificates'])
        assert_equal('checker upload concorrente recusado', response.status_code, 409)


def test_checker_validation_errors(tmpdir, paths):
    invalid_json = tmpdir / 'invalid_all_grades.json'
    invalid_json.write_text('{json invalido', encoding='utf-8')
    invalid_csv = tmpdir / 'invalid_certificados.csv'
    invalid_csv.write_text('username,course_id\nalisson,course-v1:test+python1+2026\n', encoding='utf-8')

    broken_json_paths = {**paths, 'grades': invalid_json}
    try:
        validate_checker_inputs(broken_json_paths)
        raise AssertionError('json invalido deveria falhar')
    except app_module.CourseCheckerError:
        print('OK - checker upload json invalido recusado')

    broken_csv_paths = {**paths, 'certificates': invalid_csv}
    try:
        validate_checker_inputs(broken_csv_paths)
        raise AssertionError('csv invalido deveria falhar')
    except app_module.CourseCheckerError:
        print('OK - checker upload csv invalido recusado')

    try:
        validate_checker_inputs(paths, max_grades_mb=0, max_certificates_mb=20)
        raise AssertionError('arquivo acima do limite deveria falhar')
    except app_module.CourseCheckerError:
        print('OK - checker upload limite recusado')


def test_upload_failure_keeps_status(client, report_path):
    temp_wrapper = make_temp_wrapper()

    def fake_raise(conn, path, **kwargs):
        raise RuntimeError('falha forçada')

    with patched(app_module,
                 require_operational_admin=lambda: (ADMIN, None),
                 conectar_db=lambda: type('Conn', (), {'close': lambda self: None})(),
                 upload_consumo_bloqueado=lambda *chaves: False,
                 registrar_tentativa_upload_consumo=lambda *chaves: None,
                 consumo_upload_lock=lambda conn: True,
                 get_latest_active_run=lambda conn: None,
                 importar_relatorio_checker_xlsx=fake_raise,
                 consumo_upload_unlock=lambda conn: None,
                 tempfile=type('TempModule', (), {'NamedTemporaryFile': staticmethod(lambda *args, **kwargs: temp_wrapper)})):
        with patched(app_module.app.logger, exception=lambda *args, **kwargs: None):
            response = post_upload(client, report_path)
        assert_equal('upload com falha retorna 500', response.status_code, 500)
        assert_true('upload com falha remove temporario', not Path(temp_wrapper.name).exists())


def test_status_endpoint(client):
    with patched(app_module,
                 require_roles=lambda *roles: (ADMIN, None),
                 conectar_db=lambda: type('Conn', (), {'close': lambda self: None})(),
                 get_latest_active_run=lambda conn: {'id': 99, 'status': 'running', 'started_at': '2026-06-10T10:00:00', 'finished_at': None, 'error_message': None, 'source_type': 'checker_report_xlsx', 'triggered_by_user_id': 1, 'source_files_info': {'arquivoOriginal': 'relatorio_final.xlsx'}, 'warnings': ['Processando']},
                 get_latest_successful_run=lambda conn: {'id': 88, 'status': 'success', 'started_at': '2026-06-09T10:00:00', 'finished_at': '2026-06-09T10:10:00', 'error_message': None, 'source_type': 'checker_report_xlsx', 'triggered_by_user_id': 1, 'source_files_info': {'arquivoOriginal': 'relatorio_final.xlsx'}, 'warnings': ['Anterior ok']},
                 get_run_counts=lambda conn, run_id: {'students': 12, 'courses': 34} if run_id == 99 else {'students': 11, 'courses': 33},
                 list_recent_runs=lambda conn, limit=20: []):
        response = client.get('/api/consumo/atualizacao/status')
        data = response.get_json()
        assert_equal('status endpoint 200', response.status_code, 200)
        assert_equal('status atual running', data['status'], 'running')
        assert_equal('status quantidade alunos', data['quantidadeAlunos'], 12)
        assert_equal('status quantidade cursos', data['quantidadeCursos'], 34)
        assert_true('status mostra execucao atual', data['execucaoAtual'] is not None)
        assert_true('status mostra ultima sucesso', data['ultimaAtualizacaoBemSucedida'] is not None)

    with patched(app_module,
                 require_roles=lambda *roles: (MONITOR, None),
                 conectar_db=lambda: type('Conn', (), {'close': lambda self: None})(),
                 get_latest_active_run=lambda conn: None,
                 get_latest_successful_run=lambda conn: {'id': 88, 'status': 'success', 'started_at': '2026-06-09T10:00:00', 'finished_at': '2026-06-09T10:10:00', 'error_message': None, 'source_type': 'checker_report_xlsx', 'triggered_by_user_id': 1, 'source_files_info': {'arquivoOriginal': 'relatorio_final.xlsx'}, 'warnings': ['Aviso administrativo']},
                 get_run_counts=lambda conn, run_id: {'students': 11, 'courses': 33}):
        response = client.get('/api/consumo/atualizacao/status')
        data = response.get_json()
        assert_equal('status nao admin 200', response.status_code, 200)
        assert_equal('status nao admin oculta warnings', data['warnings'], [])
        assert_equal('status nao admin oculta warnings ultima success', data['ultimaAtualizacaoBemSucedida']['warnings'], [])


def main():
    tmpdir = BACKEND_DIR / '.upload_consumo_test_tmp'
    if tmpdir.exists():
        shutil.rmtree(tmpdir)
    tmpdir.mkdir()
    try:
        report_path = tmpdir / 'relatorio_final.xlsx'
        create_report_xlsx(report_path)
        test_payload_service(report_path)
        checker_paths = create_checker_files(tmpdir)
        test_checker_payload_service(checker_paths)
        test_checker_validation_errors(tmpdir, checker_paths)

        minimal_report = tmpdir / 'relatorio_minimo.xlsx'
        create_minimal_report(minimal_report)
        with app_module.app.test_client() as client:
            test_upload_success(client, minimal_report)
            test_checker_upload_start(client, checker_paths)
            test_checker_upload_sync_success(client, checker_paths)
            test_checker_upload_sync_error(client, checker_paths)
            test_upload_restrictions(client, minimal_report)
            test_checker_upload_restrictions(client, checker_paths)
            test_upload_failure_keeps_status(client, minimal_report)
            test_status_endpoint(client)

        print('Todos os testes de upload do consumo passaram.')
    finally:
        if tmpdir.exists():
            shutil.rmtree(tmpdir)


if __name__ == '__main__':
    main()
